import time
import zipfile
import json
import shutil
import os
import pywinauto.keyboard
import pywinauto.mouse
import csv
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series
)
from openpyxl.chart.axis import ChartLines

with open(r'inAMPNoise.json') as d:
    paths = json.load(d)['Nimble'][0]

project_location = paths['project_location']
device = paths['device']
gain = paths['gain']

unzip_path_ltspice = project_location + '\\' + device + '\\' + 'LTspice - ' + device + ' G' + gain + '.zip'
print(unzip_path_ltspice)
with zipfile.ZipFile(unzip_path_ltspice) as zip_ref:
    new_path = project_location + '\\' + device
    print(new_path)
    zip_ref.extractall(new_path)

unzip_path_nimble = project_location + '\\' + device + '\\' + 'Nimble - ' + device + ' G' + gain + '.zip'
print(unzip_path_nimble)
with zipfile.ZipFile(unzip_path_nimble) as zip_ref:
    zip_ref.extractall(new_path)

time.sleep(1)
raw_data = project_location + '\\' + device + '\\' + 'Raw Data' + '\\' + 'Individual Stage Data' + '\\' + 'Amplifier' + '\\' + 'Amplifier - Output Referred Noise.csv'
shutil.move(raw_data, project_location + device)

time.sleep(3)

project_loc = paths['project_location']
device = paths['device']
gain = paths['gain']

os.startfile(project_loc + '\\' + device + '\\' + 'Noise_Simulation.asc')
time.sleep(0.25)
pywinauto.keyboard.send_keys("%{S}")
pywinauto.keyboard.send_keys("{R}")
time.sleep(1)
pywinauto.keyboard.send_keys("%{V}")
time.sleep(1)
pywinauto.keyboard.send_keys("{V}")
time.sleep(1)
pywinauto.keyboard.send_keys("{DOWN}")
pywinauto.keyboard.send_keys("{DOWN}")
pywinauto.keyboard.send_keys("{DOWN}")
pywinauto.keyboard.send_keys("{ENTER}")
time.sleep(0.25)
pywinauto.keyboard.send_keys("^{TAB}")
time.sleep(0.25)
pywinauto.mouse.right_click(coords=(50, 300))
pywinauto.keyboard.send_keys("{TAB}")
pywinauto.keyboard.send_keys("{TAB}")
pywinauto.keyboard.send_keys("{TAB}")
pywinauto.keyboard.send_keys("{TAB}")
pywinauto.keyboard.send_keys("{TAB}")
pywinauto.keyboard.send_keys("{SPACE}")
pywinauto.keyboard.send_keys("{ENTER}")
pywinauto.keyboard.send_keys("^TAB")
pywinauto.keyboard.send_keys("%{F}")
pywinauto.keyboard.send_keys("{E}")
pywinauto.keyboard.send_keys("{ENTER}")
pywinauto.keyboard.send_keys("%{F4}")

time.sleep(0.5)

extra_files_remove = project_loc + device
zip_remove_nimble = extra_files_remove + '\\' + 'Nimble - ' + device + ' G' + gain + '.zip'
zip_remove_ltspice = extra_files_remove + '\\' + 'LTspice - ' + device + ' G' + gain + '.zip'

time.sleep(2)

path_file = pd.read_csv(paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Output Referred Noise.csv')
path_file.to_excel(paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Output Referred Noise.xlsx',index=None, header=True)

excel_path = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Output Referred Noise.xlsx'

file = openpyxl.load_workbook(excel_path)
sheet_obj = file.active
sheet_obj.delete_cols(3)
time.sleep(0.25)
sheet_obj.delete_cols(3)
time.sleep(0.25)
sheet_obj.delete_cols(3)
time.sleep(0.25)
sheet_obj.delete_cols(3)
file.save(excel_path)

text_file = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Noise_Simulation.txt'
output_file = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Noise_Simulation.xlsx'
gain_textfile = paths['project_location'] + '\\' + paths['device'] + '\\' + paths['device'] + ' G' + paths[
    'gain'] + '.txt'

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

with open(text_file, 'r') as data:
    reader = csv.reader(data, delimiter='\t')
    for row in reader:
        ws.append(row)

wb.save(output_file)

path1 = output_file
path2 = excel_path

wb1 = openpyxl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

wb2 = openpyxl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)

for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value

wb2.save(path2)

xl = openpyxl.load_workbook(excel_path)
sheet1 = xl['Sheet']
sheet2 = xl['Sheet1']

columnA = []

for i in range(1, 1000, 1):
    columnA.append(sheet1.cell(row=i, column=1).value)

for i in range(1, 1000, 1):

    for i in range(1, 1000, 1):
        sheet2.cell(row=i, column=4).value = columnA[i - 1]

columnB = []

for i in range(1, 1000, 1):
    columnB.append(sheet1.cell(row=i, column=2).value)

for i in range(1, 1000, 1):

    for i in range(1, 1000, 1):
        sheet2.cell(row=i, column=5).value = columnB[i - 1]

xl.save(excel_path)

xr = pd.read_excel(excel_path)

xr.rename(columns={xr.columns[0]: 'Frequency (Hz)'}, inplace=True)
xr.rename(columns={xr.columns[1]: 'Total (nV/rt(Hz))'}, inplace=True)
xr.rename(columns={xr.columns[2]: ''}, inplace=True)
xr.rename(columns={xr.columns[3]: 'Ltspice Freq'}, inplace=True)
xr.rename(columns={xr.columns[4]: 'Ltspice V(onoise)'}, inplace=True)
# rename_column.drop(columns={rename_column.columns[0]}, axis=1, inplace=True)

xr['Total (nV/rt(Hz))'] = xr['Total (nV/rt(Hz))'].apply(lambda x: x * 1e9)
xr['Ltspice V(onoise)'] = xr['Ltspice V(onoise)'].apply(lambda x: x * 1e9)

xr.to_excel(excel_path, sheet_name='G ' + paths['gain'])

time.sleep(3)

extra_files_remove = paths['project_location'] + paths['device']
zip_remove_nimble = extra_files_remove + '\\' + 'Nimble - ' + paths['device'] + ' G' + paths['gain'] + '.zip'
zip_remove_ltspice = extra_files_remove + '\\' + 'LTspice - ' + paths['device'] + ' G' + paths['gain'] + '.zip'
os.remove(extra_files_remove + '\\' + 'AC_Simulation.asc')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.asc')
os.remove(extra_files_remove + '\\' + 'Transient_Simulation.asc')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.log')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.raw')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.txt')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.xlsx')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.op.raw')
os.remove(extra_files_remove + '\\' + 'Amplifier - Output Referred Noise.csv')
os.remove(zip_remove_nimble)
os.remove(zip_remove_ltspice)
shutil.rmtree(extra_files_remove + '\\' + 'Raw Data')

time.sleep(2)

with open(r'inAMPNoise.json') as d:
    paths = json.load(d)['Nimble'][0]

workbook_path = (paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Output Referred Noise.xlsx')
workbook = load_workbook(workbook_path)
sheet = workbook.active

x_nimble = Reference(sheet, min_col=3, min_row=2, max_row=1000)
y_nimble = Reference(sheet, min_col=2, min_row=2, max_row=1000)
x_ltspice = Reference(sheet, min_col=6, min_row=2, max_row=1000)
y_ltspice = Reference(sheet, min_col=5, min_row=2, max_row=1000)
# x_datasheet = Reference(sheet, min_col=6, min_row=2, max_row=1000)
# y_datasheet = Reference(sheet, min_col=5, min_row=2, max_row=1000)

series_voltage = Series(x_nimble, y_nimble, title_from_data=False, title="Nimble")
series_freq = Series(x_ltspice, y_ltspice, title_from_data=False, title="LTspice")
# series_mag = Series(x_datasheet, y_datasheet,
#                      title_from_data=False, title="Datasheet")

# Chart type
chart = ScatterChart()
chart.series.append(series_voltage)
chart.series.append(series_freq)
# chart.series.append(series_mag)
chart.x_axis.scaling.logBase = 10
chart.y_axis.scaling.logBase = 10
chart.x_axis.number_format = '0.00E+00'
chart.x_axis.minorGridlines = ChartLines()
chart.y_axis.minorGridlines = ChartLines()

chart.x_axis.scaling.min = paths['x_axis_min']
chart.y_axis.scaling.min = paths['y_axis_min']
chart.x_axis.scaling.max = paths['x_axis_max']
chart.y_axis.scaling.max = paths['y_axis_max']
chart.x_axis.tickLblPos = "low"
# chart.x_axis.tickLblSkip = 3

chart.title = None
chart.x_axis.title = 'Frequency (Hz)'
chart.y_axis.title = 'Voltage Noise Density (nV???Hz)'
chart.legend.position = 'r'

sheet.add_chart(chart, 'K3')
workbook.save(workbook_path)

time.sleep(1)

rename_noise_excel = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Output Referred Noise.xlsx'
new_noise_excel = paths['project_location'] + '\\' + paths['device'] + '\\' + paths['device'] + '.xlsx'
os.rename(rename_noise_excel, new_noise_excel)

